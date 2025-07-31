import openpyxl as xl
import sys
import os

from tkinter import Tk     # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename



# Convert/Download Bank Transactions into excel or manual input into excel
# Some Cells are Allocated to specify the output (img/graph/etc and location - inside main or elsewhere)

# Download Monthly Transaction then make a conclusive analysis -> Chart/Ledger/BOOM BOOM


Tk().withdraw()

global filePath
filePath = askopenfilename() # show an "Open" dialog box and return the path to the selected file

print(filePath)

# Input 

class DataReader:
    def __init__(self) -> None:
        self.dataLocation = filePath + "newData"
        self.dataList = {} # Title : {Date / Account / Reason / Money} - key (top/head cell) Data (rest of the column/row)
        self.setInstances()

    def requestExcel(self, file) -> dict:
        workbook = xl.load_workbook(self.dataLocation + "\\" + file)
        worksheets = workbook.sheetnames
        worksheetData = {}

        for worksheetTitle in worksheets:
            worksheet = workbook[worksheetTitle]
            data = workbook.copy_worksheet(worksheet)
            worksheetData[file + " " + worksheetTitle] = data

        workbook.close()
        return worksheetData
        
    def setInstances(self):
        files = os.listdir(self.dataLocation)
        for file in files:
            if not file.endswith("xlsx"):
                continue
            
            data = self.requestExcel(file)
            self.dataList.update(data)
            # move file to usedData or delete

    def getData(self) -> dict:
        return self.dataList



# Algorithm




class DataVisualiser:
    def __init__(self, data) -> None:
        self.data = data

    def visualiseAllData(self):
        titles = self.data.keys()
        for title in titles:
            self.visualiseData(title, self.data[title])
        # Title : {Date / Place / Reason / Money} 
        # - key (top/head cell) Data (rest of the column/row)

    def visualiseData(self, title, data):
        pass



# --graph of profit/loss in excel

# --predict financial outcomes (separate worksheet)





# Output

class DataOutput:
    def __init__(self, data, adds, file="main.xlsx") -> None:
        self.outputFile = filePath + file
        self.data = data
        self.additional = adds

    def setOutputFile(self, file) -> None:
        self.outputFile = file

    def addAllData(self) -> None:
        titles = self.data.keys()
        for title in titles:
            self.addData(title, self.data[title])

    def addData(self, title, data) -> None:
        workbook = xl.load_workbook(self.outputFile)
        sheet = workbook.create_sheet(title=title)

        for row in data:
            for cell in row:
                sheet[cell.coordinate].value = cell.value

        workbook.save(self.outputFile)

# --add into existing financial portfolio



if __name__ == "__main__":
    reader = DataReader()
    data = reader.getData()
    print(data)
#     visualiser = DataVisualiser(data)
#     visuals = visualiser.visualiseAllData() # Graphs or other additional data
    
#     output = DataOutput(data=data, adds=visuals)
#     output.addAllData()
    

# print(os.listdir(".\\coding\\clean_code\\python\\money\\newData"))